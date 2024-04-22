import math
import re
import unicodedata
from re import compile
import random
import string
import os.path
from openpyxl import *
import datetime
from typing import Optional, List
from robot.api.logger import console
from robot.libraries.BuiltIn import BuiltIn
from robot.model.testsuite import TestSuite
from robot.running.model import TestCase
from robot.utils.dotdict import DotDict
from robot.libraries.BuiltIn import BuiltIn
from robot.api.deco import keyword
from enum import Enum, auto
from XrayInterfaceAPI import *

#pdf
from fpdf import FPDF, HTMLMixin

built_in = BuiltIn()

class PabotOpt(Enum):
    Equal = auto()
    Binary = auto()
    Atomic = auto()

class ControldeVariablesLibrary:
    ROBOT_LIBRARY_SCOPE = 'SUITE'
    ROBOT_LIBRARY_DOC_FORMAT = "reST"
    ROBOT_LIBRARY_VERSION = "1.6.1"
    ROBOT_LISTENER_API_VERSION = 3

    def __init__(
        self,
        archivoVariables: str = None,
        pais: str = None,
        delimitador: Optional[str] = "|",
        hoja: Optional[str] = None,
        optimize_pabot: Optional[PabotOpt] = PabotOpt.Atomic,
    ):
        self.ROBOT_LIBRARY_LISTENER = self
        self.archivoVariables=archivoVariables
        self.pais=pais
        self.delimitador=delimitador
        self.variables_globales={}
        self.datos={}
        if not self.delimitador:
            self.delimitador="|"
        self.hoja=hoja
        self.orden='Suite'
        if not self.hoja:
            self.hoja = self.pais
        self.data_table_dict = DotDict()
        self.RANDOM_STRING_PATTERN = compile(r"(?i)^#(aleatorio|random)(texto|cadena|string)(\d*)$")
        self.RANDOM_STRING_RANGE_PATTERN = compile(r"(?i)^#(aleatorio|random)(texto|cadena|string)(\d+)$")
        #self.RANDOM_STRING_PATTERN = compile(r"(?i)^#aleatorio_texto?_?\d*$")
        #self.RANDOM_STRING_PATTERN2 = compile(r"(?i)^#[(aleatorio)|(random)]_texto?_?\d+$")
        self.RANDOM_NUMERO_PATTERN = compile(r"(?i)^#(aleatorio|random)numero(\d*)$")
        self.RANDOM_NUMERO_RANGE_PATTERN = compile(r"(?i)^#(aleatorio|random)numero(\d+)$")
        self.RANDOM_ALFANUMERICO_PATTERN = compile(r"(?i)^#(aleatorio|random)alfanumerico(\d*)$")
        self.RANDOM_ALFANUMERICO_RANGE_PATTERN = compile(r"(?i)^#(aleatorio|random)alfanumerico(\d+)$")
        self.RANDOM_MAIL_PATTERN = compile(r"(?i)^#(aleatorio|random)(e?mail|correo)(\d*)$")
        self.RANDOM_MAIL_RANGE_PATTERN = compile(r"(?i)^#(aleatorio|random)(e?mail|correo)(\d+)$")
        self.RANDOM_NIT_PATTERN = compile(r"(?i)^#(aleatorio|random)nit(\d*)$")
        self.RANDOM_NIT_RANGE_PATTERN = compile(r"(?i)^#(aleatorio|random)nit(\d+)$")
        self.RANDOM_NIT_TEXTO_PATTERN = compile(r"(?i)^#(aleatorio|random)nit(texto|cadena|string)(\d*)$")
        self.RANDOM_NIT_TEXTO_RANGE_PATTERN = compile(r"(?i)^#(aleatorio|random)nit(texto|cadena|string)(\d+)$")
        self.RANDOM_NIT_SV_PATTERN = compile(r"(?i)^#(aleatorio|random)nitsv(\d*)$")
        self.RANDOM_NIT_SV_RANGE_PATTERN = compile(r"(?i)^#(aleatorio|random)nitsv(\d+)$")
        self.optimize_pabot=optimize_pabot;

    
    def _start_suite(self, suite: TestSuite, *_):
        try:
            self.suite_name = suite.longname
            self.file_array=[]            
            self.template_tests = suite.tests
            self.original_tests = suite.tests
            self.suite_source = suite.source
            console(f"Suite: {self.suite_name} - pais: {self.pais} - hoja: {self.hoja} - # tests: {len(suite.tests)}")
            self.datos=self.obteniendo_informacion()
            self._read_value_from_datos(self.variables_globales,1)
            try:
                self.orden = built_in.replace_variables('${orden_a_ejecutar}')
            except Exception as e:
                self.orden = 'Suite'
            test_list = list()
            if self.is_same_values(self.orden, 'ListadodeExcel'):
                test_list = self._get_list_test_cases_by_excel(suite)
            elif self.is_same_values(self.orden, 'Regresivo'):
                test_list = self._get_list_test_cases(suite)
            elif self.is_same_values(self.orden, 'Suite'):
                test_list = self._get_list_test_cases(suite)
            else:
                console('',newline=True)
                console('No fue posible obtener el listado de Pruebas, únicos posibles valores: Suite o ListadodeExcel en la variable orden_a_ejecutar')
            if self._handle_pabot(test_list):
                suite.tests = []
            else:
                suite.tests = test_list
            self.update_suite_summary(self.suite_name,None,1)
            console('Suite establecido.', newline=True)
        except Exception as exception:
            console(f"Error encontrado: {exception}")
        console('------------------------------------------------------------------------------',newline=True)


    def _handle_pabot(self, test_list):
        if (
            not built_in.get_variable_value("${PABOTQUEUEINDEX}")
        ):
            return
        console(f'manejado por pabot')
        pabot_process_count = int(built_in.get_variable_value("${PABOTNUMBEROFPROCESSES}"))
        if not pabot_process_count:
            console('No se tiene un número de proceso')
            return
        try:
            from pabot.pabotlib import Remote  # type: ignore
        except ImportError as e:
            console(f'Error importando PabotLib {e}.')
            return
        pabotlib_url = built_in.get_variable_value("${PABOTLIBURI}")
        try:
            pabotlib = Remote(pabotlib_url)
            if not pabotlib:
                raise ConnectionError
            self._create_pabot_queue(pabot_process_count, pabotlib, test_list)
        except (RuntimeError, ConnectionError) as e:
            console(f'Ejecución sin --testlevelsplit, o sin usar --pabotlib, error: {e}, pabotlib_url: {pabotlib_url}')
            return
        pabotlib.run_keyword("ignore_execution", [built_in.get_variable_value("${CALLER_ID}")], {})
        console(f'pabot lib uri: {pabotlib_url}.')
        return True


    def _create_pabot_queue(self, pabot_process_count, pabotlib, test_list):
        pabot_opt = self.optimize_pabot
        console(f'create Pabot queue, Pabot type: {pabot_opt}')
        if pabot_opt == PabotOpt.Atomic:
            self._add_test_to_pabot_queue(pabotlib, test_list)
        else:
            if pabot_opt == PabotOpt.Binary:
                console("nada2")
                partitioner = self.binary_partition_test_list(test_list, pabot_process_count)
                console(f"nada, partitioner: {partitioner}")
            else:
                console(f'Longitud de casos: {len(test_list)}.')
                partitioner = self.equally_partition_test_list(test_list, pabot_process_count)
                console(f'partitioner: {partitioner}.,.')
            for item in partitioner:
                console(f'item: {item}.')
                for item2 in item:
                    console(f'item2: {item2}.')
            console(f'iniciar con el siguiente valor en partitioner: -{partitioner}-')
            console(f'partitioner: -{partitioner(test_list, pabot_process_count)}-')
            for process_test_list in partitioner(test_list, pabot_process_count):
                console(f'process_test_list: {process_test_list}.')
                self._add_test_list_to_pabot_queue(pabotlib, process_test_list)


    def equally_partition_test_list(self,test_list: List, fraction_count: int):
        quotient, remainder = divmod(len(test_list), fraction_count)
        console(f'quotient, remainder: {quotient}, {remainder}.')
        return [
            test_list[i * quotient + min(i, remainder) : (i + 1) * quotient + min(i + 1, remainder)]
            for i in range(fraction_count)
        ]


    def binary_partition_test_list(self,test_list: List, process_count: int):
        console('ingreso a binary_partition_test_list')
        fractions = self.equally_partition_test_list(test_list, process_count)
        console(f'binary_partition_test_list() - fractions {fractions}')
        return_list = list()
        for i in range(int(math.sqrt(len(test_list) // process_count))):
            first, second = self._partition_second_half(fractions)
            return_list.extend(first)
            fractions = second
        return_list.extend(fractions)
        return [test_name for test_name in return_list if test_name]


    def _add_test_to_pabot_queue(self, pabotlib, test_list):
        for test in test_list:
            console('pruebaxyz')
            pabotlib.run_keyword(
                "add_suite_to_execution_queue",
                [self.suite_name, [f"DYNAMICTEST:{self.suite_name}.{test.name}"]],
                {},
            )


    def _add_test_list_to_pabot_queue(self, pabotlib, test_list):
        console('start _add_test_list_to_pabot_queue()')
        test_names = [f"{self.suite_name}.{test.name}" for test in test_list]
        console(f'test_names: {test_names}')
        pabot_string = '|'.join(
            [name.replace('\\', '\\\\').replace('|', '\\|') for name in test_names]
        )
        console(f'pabot_string: {pabot_string}')
        pabotlib.run_keyword(
            "add_suite_to_execution_queue",
            [self.suite_name, [f"DYNAMICTESTS:{pabot_string}"]],
            {},
        )
    

    def _partition_second_half(self,fractions):
        first = fractions[: len(fractions) // 2]
        second = list()
        for sub_list in fractions[len(fractions) // 2 :]:
            sub_sub_list = self.equally_partition_test_list(sub_list, 2)
            console(f'sub_sub_list: {sub_sub_list}')
            second.extend(sub_sub_list)
        return first, second


    def equally_partition_test_list(self,test_list: List, fraction_count: int):
        quotient, remainder = divmod(len(test_list), fraction_count)
        return [
            test_list[i * quotient + min(i, remainder) : (i + 1) * quotient + min(i + 1, remainder)]
            for i in range(fraction_count)
        ]


    def _get_list_test_cases_by_excel(self, suite: TestSuite):
        temp_test_list = list()
        temp_datos = self.datos
        temp_test2=self.template_tests
        console('',newline=True)
        #console(f'datos: .{temp_datos}.')
        console("Encontrado by excel:",newline=False)
        for elemento in temp_datos:
            for modulo in elemento.keys():
                if self.is_same_values(str(modulo),'General'):
                    console('general (continuando)..., ', newline=False)
                    continue
                var_agregar_tc = False
                for var_test in temp_test2:
                    self.template_test=suite.tests[temp_test2.index(var_test)]
                    for test in self.template_test.tags:
                        if self.is_same_values(test,modulo):
                            console(f' .{str(test)}.== .{str(modulo)}. bloque {(temp_datos.index(elemento)+1)} con [{len(elemento.get(modulo))}]*,',newline=False)
                            for n in range(0,len(elemento.get(modulo))):
                                self._create_test_from_template(True, (temp_datos.index(elemento) + 1), (n+1), len(elemento.get(modulo)))
                                temp_test_list.append(self.test)
                            var_agregar_tc = True
                    if var_agregar_tc:
                        ######temp_datos.pop(modulo)
                        break
                if var_agregar_tc == False:
                    console(f' .{str(modulo)}. con [0],',newline=False)
        console('',newline=True)
        return temp_test_list


    def _get_list_test_cases(self, suite: TestSuite):
        temp_test_list = list()
        temp_datos = self.datos
        temp_test2=self.template_tests
        console("Encontrado",newline=False)
        for var_test in temp_test2:
            self.template_test=suite.tests[temp_test2.index(var_test)]
            #####console(f'template_test: {self.template_test.tags}')
            var_agregar_tc = False
            #console(f'----Test: {self.template_test.name}')
            for elemento in temp_datos:
                for modulo in elemento.keys():
                    #console(f' modulo: {modulo}',newline=False)
                    if self.is_same_values(str(modulo),'General'):
                        console('general (continuando)..., ', newline=False)
                        continue
                    #console("valor 1: {v1} - valor 2: {v2}".format(v1=var_modmulo,v2=self.template_test.name))
                    for test in self.template_test.tags:
                        if self.is_same_values(test,modulo):
                            console(f' .{str(test)}.== .{str(modulo)}. con [{len(elemento.get(modulo))}]*,',newline=False)
                            for n in range(0,len(elemento.get(modulo))):
                                self._create_test_from_template(False, 0, (n+1), len(elemento.get(modulo)))
                                temp_test_list.append(self.test)
                            var_agregar_tc = True
                    if var_agregar_tc:
                        ######temp_datos.pop(modulo)
                        break
                if var_agregar_tc:
                    break
            if var_agregar_tc == False:
                console(f' .{self.template_test.tags}. con [1],',newline=False)
                self._create_test_from_template(False, 0, 1,1)
                var_agregar_tc = True
                temp_test_list.append(self.test)
                #temp_test2.pop(temp_test2.index(var_test))
        console('',newline=True)
        return temp_test_list


    def obteniendo_informacion(self):
        datos=[]
        path_archivo = self.archivoVariables
        nombre_hoja = self.hoja
        hojaDelPais = load_workbook(filename=os.path.abspath(path_archivo), data_only=True)
        if nombre_hoja not in hojaDelPais.sheetnames:
            console(f'Hoja: {nombre_hoja} no encontrado en el archivo {path_archivo}')
            console('',newline=True)
            return datos
        sheet   = hojaDelPais[nombre_hoja]
        rowmax = 0
        colmax = 0
        rowmin = 0
        colmin = 0
        console('módulos: ', newline=False)
        try:
            colmin = sheet.min_column
            rowmin = sheet.min_row
            colmax=sheet.max_column
            rowmax=sheet.max_row
            existe_variables_generales=False
            if (colmax-colmin) < 1 and (rowmax-rowmin) < 1:
                console('No se encontró datos en el archivo adjunto',newline=False)
                console('',newline=True)
                return datos
            #console(" columna mínimo: {cmi} máximo: {cma} - fila mínimo: {fmi} máximo: {fma}".format(cmi=colmin,cma=colmax,fmi=rowmin,fma=rowmax),newline=False)
            for i in range(colmin,colmax):
                if i == 0:
                    continue
                for j in range(rowmin,(rowmax+1)):
                    #print(f"fila: {j}")
                    if j == 0:
                        continue
                    if sheet.cell(row=j, column=i).value == None:
                        continue
                    console(" {v},".format(v=sheet.cell(row=j, column=i).value),newline=False)
                    contador=0
                    for x1 in range(j+1,(rowmax+1)):
                        if sheet.cell(row=x1, column=i).value != None:
                            break
                        contador = contador +1
                    if contador < 1:
                        if j < rowmax:
                            contador=rowmax-j
                        else:
                            print("Error - Los datos recibidos no tienen valores", end='')
                            continue
                    valores = []
                    elemento={}
                    for j2 in range(0, contador):
                        counter_row=j2+1
                        arg = {}
                        tiene_valores = False
                        for i2 in range(i+1, colmax+1):
                            #nombre del atributo
                            if sheet.cell(row=j, column=i2).value == None:
                                continue
                            #celda
                            if sheet.cell(row=(j+counter_row), column=i2).value == None:
                                val_celda = ''
                            else:
                                tiene_valores=True
                                val_celda = sheet.cell(row=(j+counter_row), column=i2).value
                            column_name = sheet.cell(row=j, column=i2).value
                            column_name = column_name.lower()
                            #array_val_celda = val_celda.split(sep=separador_celda)
                            #arg[sheet.cell(row=j, column=i2).value] = array_val_celda
                            arg[column_name] = val_celda
                            #console(" columna: {c} valor: {v}.".format(c=column_name,v=arg[column_name]),newline=False)
                        if not len(arg) == 0:
                            if tiene_valores:
                                valores.append(arg)
                    funcionalidad_tmp = sheet.cell(row=j, column=i).value
                    funcionalidad_tmp = funcionalidad_tmp.lower()
                    if self.is_same_values(funcionalidad_tmp,'General'):
                        if existe_variables_generales == False:
                            self.variables_globales=valores[0]
                            existe_variables_generales=True
                    else:
                        elemento[funcionalidad_tmp] = valores
                        datos.append(elemento)
                break
            else:
                print("No contiene valores el archivo recibido!")
        except ValueError:
            print("--------Error: {e}".format(e=ValueError))
        console('',newline=True)
        return datos


    def _create_test_from_template(self,
        tipo: bool=False, #False Suite, #True por Excel
        bloque: int=0,
        contador: int=0,
        max_num_test_a_crear: Optional[int]=1,
    ):
        complemento_nombre=''
        if tipo:
            complemento_nombre= '_Bloque_' + str(bloque)
        if max_num_test_a_crear > 1:
                complemento_nombre= complemento_nombre + '_Ejecución_' + str(contador)
        self.test = TestCase(
            name=self.template_test.name+complemento_nombre,
            doc=self.template_test.doc,
            tags=self.template_test.tags,
            template=self.template_test.template,
            lineno=self.template_test.lineno,
            timeout=self.template_test.timeout,
        )
        self.test.setup=self.template_test.setup
        self.test.teardown=self.template_test.teardown
        self.test.body=self.template_test.body
        self.test.parent = self.template_test.parent


    def _start_test(self, test: TestCase, *_):
        console("",newline=True)
        console(f"Test a ejecutar:{test.name}, tags: {test.tags}:")
        self._fill_variables_for_test_case(test)
        var_test = str(test.name)        
        self._start_testcase_evidence(var_test)
        console('Listo para iniciar', newline=True)
        
        
    def _start_testcase_evidence(self,testname:str=''):
        try:
            array_body = []
            body_test_dic = {}
            try:
                is_save_by_test = BuiltIn().get_variable_value('${saveBYTEST}')
                if is_save_by_test=='False':
                    array_body = built_in.get_variable_value('${bodyHTML2}')
                else:
                    built_in.set_suite_variable('${tablaResumenBodyPDF}','')
                    built_in.set_suite_variable('${numTCSummary}',0)
            except Exception as exception:
                array_body = []
            var_body_new = """<h3 align="left"><font color="black">Test: {tc}</font></h3>
            <p>&nbsp;</p>"""
            var_body_new = var_body_new.format(tc=testname)
            body_test_dic['Titulo'] = var_body_new
            body_test_dic['Body'] = ''
            body_test_dic['Resultado'] = ''
            array_body.append(body_test_dic)
            built_in.set_suite_variable('${bodyHTML2}',array_body)
            built_in.set_suite_variable('${tmpFunct}','0')
            built_in.set_suite_variable('${file_arrays}',[])
            #console(f'array_body: {array_body}')
        except Exception as exception:
            console(f'Exception in _start_testcase_evidence(), excepción: {exception}', newline=False)


    def _fill_variables_for_test_case(self,test):
        try:
            console('Obteniendo variables: ', newline=False)
            if re.search(r'.*bloque_(\d)+.*',str(test.name).lower()):
                for tag in test.tags:
                    x = re.search(r'(?<=bloque_)\d+',str(test.name).lower())
                    bloque=x.group(0)
                    if not self.datos[int(bloque)-1].get(str(tag).lower()):
                        continue
                    if re.search(r'.*ejecución_(\d)+$',str(test.name).lower()):
                        x = re.search(r'(?<=ejecución_)\d+',str(test.name).lower())
                        console(f'módulo {str(tag).lower()} - bloque: {bloque} - posición del mapa: {x.group(0)}.', newline=False)
                        self._read_value_from_datos(self.datos[int(bloque)-1].get(str(tag).lower())[(int(x.group(0))-1)],2,tag,x.group(0))
                        break
                    else:
                        console(f'módulo {str(tag).lower()} - bloque: {bloque} - posición del mapa: 1.', newline=False)
                        self._read_value_from_datos(self.datos[int(bloque)-1].get(str(tag).lower())[0],2,tag)
                        break
            else:
                is_saved=False
                for elemento in self.datos:
                    if not is_saved:
                        for modulo in elemento.keys():
                            for tag in test.tags:
                                if not self.is_same_values(str(modulo).lower(),str(tag).lower()):
                                    continue
                                if re.search(r'.*ejecución_(\d)+$',str(test.name).lower()):
                                    x = re.search(r'(?<=ejecución_)\d+',str(test.name).lower())
                                    console(f'módulo {str(tag).lower()} - posición del mapa: {x.group(0)}.', newline=False)
                                    self._read_value_from_datos(elemento.get(modulo)[(int(x.group(0))-1)],2,tag,x.group(0))
                                    is_saved=True
                                    break
                                else:
                                    console(f'módulo {str(tag).lower()} - posición del mapa: 1.', newline=False)
                                    self._read_value_from_datos(elemento.get(modulo)[0],2,tag)
                                    is_saved=True
                                    break
                if not is_saved:
                    console('no tiene variables', newline=False)
        except Exception as e:
            console(f'Exception in _fill_variables_for_test_case(), excepción: {e}', newline=False)
        console('',newline=True)


    def _read_value_from_datos(self, 
        argumentos: dict,
        tipo: int,
        tc: Optional[str]='',
        ejecucion: Optional[str]='',
    ):
        for argumento in argumentos.keys():
            variable_value = argumentos.get(argumento)
            #console(f' valor: {variable_value}', newline=False)
            variable_value = self._is_random(str(variable_value), str(argumento))
            variable_string = '${'+str(argumento)+'}'
            variable_string = f"${variable_string[1:]}"
            #console(f'variable: {variable_string} - nvalor: {variable_value},', newline=False)
            if tipo == 1: #1 - Suite
                built_in.set_suite_variable(variable_string,variable_value)
            elif tipo == 2: #2 - Test Case
                built_in.set_test_variable(variable_string,variable_value)
        if tipo == 2:
            built_in.set_test_variable('${ImagenID_TC}',tc)
            built_in.set_test_variable('${EjecucionID_TC}', ejecucion)
        elif tipo == 1:
            built_in.set_suite_variable('${numTCSummary}',0)
            built_in.set_suite_variable('${TCSummaryList}',[])


    def return_values_from_array(self, values):
        return str(values).split(self.delimitador)


    def _is_array(self, value):
        return str(value).__contains__(self.delimitador)


    def is_same_values(self, first: str, second: str):
        return self._get_normalized_keyword(first) == self._get_normalized_keyword(second)


    def _get_normalized_keyword(self, keyword: str):
        return self._strip_accents(keyword).lower().replace(" ", "").replace("_", "")


    def _strip_accents(self, keyword):
        return ''.join(c for c in unicodedata.normalize('NFD', keyword)
                   if unicodedata.category(c) != 'Mn')


    def _get_first_random_number(self):
        primer_valor=''
        while True:
            primer_valor = ''.join(random.choice(string.digits) for i in range(1))
            #console(f'numero: {primer_valor}')
            if primer_valor != '0':
                break
        return primer_valor
    
    
    def _is_random(self,valor, key):
        valor_str = ''
        try:
            if self.RANDOM_STRING_PATTERN.match(self._get_normalized_keyword(str(valor))):
                if self.RANDOM_STRING_RANGE_PATTERN.match(self._get_normalized_keyword(valor)):
                    numero= re.search(r'(\d)+$',str(valor))
                    valor_str = ''.join(random.choice(string.ascii_lowercase) for i in range(int(numero.group(0))))
                else:
                    valor_str = ''.join(random.choice(string.ascii_lowercase) for i in range(10))
            elif self.RANDOM_ALFANUMERICO_PATTERN.match(self._get_normalized_keyword(str(valor))):
                if self.RANDOM_ALFANUMERICO_RANGE_PATTERN.match(self._get_normalized_keyword(str(valor))):
                    numero= re.search(r'(\d)+$',str(valor))
                    while len(valor_str) < int(numero.group(0)):
                        valor_str = valor_str + ''.join(random.choice(string.ascii_lowercase))
                        if len(valor_str) < int(numero.group(0)):
                            valor_str = valor_str + ''.join(random.choice(string.digits))
                else:
                    while len(valor_str) < 10:
                        valor_str = valor_str + ''.join(random.choice(string.ascii_lowercase))
                        if len(valor_str) < 10:
                            valor_str = valor_str + ''.join(random.choice(string.digits))
            elif self.RANDOM_MAIL_PATTERN.match(self._get_normalized_keyword(str(valor))):
                if self.RANDOM_MAIL_RANGE_PATTERN.match(self._get_normalized_keyword(str(valor))):
                    numero= re.search(r'(\d)+$',str(valor))
                    valor_str = ''.join(random.choice(string.ascii_lowercase+string.digits) for i in range(int(numero.group(0))))
                else:
                    valor_str = ''.join(random.choice(string.ascii_lowercase+string.digits) for i in range(10))
                valor_str = valor_str + '@mail.com'
            elif self.RANDOM_NUMERO_PATTERN.match(self._get_normalized_keyword(str(valor))):
                if self.RANDOM_NUMERO_RANGE_PATTERN.match(self._get_normalized_keyword(str(valor))):
                    numero= re.search(r'(\d)+$',str(valor))
                    primer_digito=self._get_first_random_number()
                    valor_str = primer_digito + ''.join(random.choice(string.digits) for i in range(int(numero.group(0)) -1))
                else:
                    primer_digito=self._get_first_random_number()
                    valor_str = primer_digito + ''.join(random.choice(string.digits) for i in range(9))
            elif self.RANDOM_NIT_PATTERN.match(self._get_normalized_keyword(str(valor))):
                if self.RANDOM_NIT_RANGE_PATTERN.match(self._get_normalized_keyword(str(valor))):
                    numero= re.search(r'(\d)+$',str(valor))
                    valor_str=self._get_first_random_number()
                    while len(valor_str) < (int(numero.group(0))-2):
                        valor_str = valor_str + ''.join(random.choice(string.digits))
                    valor_str = valor_str + '-' + ''.join(random.choice(string.digits) for i in range(1))
                else:
                    valor_str=self._get_first_random_number()
                    while len(valor_str) < 8:
                        valor_str = valor_str + ''.join(random.choice(string.digits))
                    valor_str = valor_str + '-' + ''.join(random.choice(string.digits) for i in range(1))
            elif self.RANDOM_NIT_TEXTO_PATTERN.match(self._get_normalized_keyword(str(valor))):
                if self.RANDOM_NIT_TEXTO_RANGE_PATTERN.match(self._get_normalized_keyword(str(valor))):
                    numero= re.search(r'(\d)+$',str(valor))
                    valor_str=self._get_first_random_number()
                    while len(valor_str) < (int(numero.group(0))-2):
                        valor_str = valor_str + ''.join(random.choice(string.digits))
                    valor_str = valor_str + '-' + ''.join(random.choice(string.ascii_lowercase) for i in range(1))
                else:
                    valor_str=self._get_first_random_number()
                    while len(valor_str) < 8:
                        valor_str = valor_str + ''.join(random.choice(string.digits))
                    valor_str = valor_str + '-' + ''.join(random.choice(string.ascii_lowercase) for i in range(1))
            elif self.RANDOM_NIT_SV_PATTERN.match(self._get_normalized_keyword(str(valor))):
                if self.RANDOM_NIT_SV_RANGE_PATTERN.match(self._get_normalized_keyword(str(valor))):
                    numero= re.search(r'(\d)+$',str(valor))
                    valor_str=self._get_first_random_number()
                    var_size = int(numero.group(0))
                    while len(valor_str) < var_size:
                        valor_str = valor_str + ''.join(random.choice(string.digits))
                    valor_str = valor_str + '-' + self._get_first_random_number()
                    var_size = var_size + 7
                    while len(valor_str) < var_size:
                        valor_str = valor_str + ''.join(random.choice(string.digits))
                    valor_str = valor_str + '-' + self._get_first_random_number()
                    var_size = var_size + 4
                    while len(valor_str) < var_size:
                        valor_str = valor_str + ''.join(random.choice(string.digits))
                    valor_str = valor_str + '-' + ''.join(random.choice(string.digits) for i in range(1))
                else:
                    valor_str=self._get_first_random_number()
                    while len(valor_str) < 4:
                        valor_str = valor_str + ''.join(random.choice(string.digits))
                    valor_str = valor_str + '-' + self._get_first_random_number()
                    while len(valor_str) < 11:
                        valor_str = valor_str + ''.join(random.choice(string.digits))
                    valor_str = valor_str + '-' + self._get_first_random_number()
                    while len(valor_str) < 15:
                        valor_str = valor_str + ''.join(random.choice(string.digits))
                    valor_str = valor_str + '-' + self._get_first_random_number()
            else:
                valor_str=valor
            if not self.is_same_values(valor_str,valor):
                console(f' {key} nuevo valor: {valor_str},',newline=False)
        except Exception as e:
            console(f'Se obtuvo un error al generar el valor random del siguiente valor: {valor}, siguiente excepcion: {e}.', newline=False)
        return valor_str


    def _get_value(self, valor_atributo):
        resultado = ''
        try:
            #console(f'_get_value(): valor_atributo= {valor_atributo}')
            if valor_atributo:
                listado = re.findall(r'[0-9]+',str(valor_atributo))
                for valor_atributo in listado:
                    #console(f'valor_atributo 2: {valor_atributo}')
                    resultado += valor_atributo
                numero_del_tc = 0
                if len(resultado) >= 1:
                    numero_del_tc = int(resultado)
                resultado = 'TC'+str(numero_del_tc)
            else:
                console(f'valor vacio: {valor_atributo}')
        except Exception as e:
            console(f'Exception in _get_value(), excepción: {e}')
        return resultado
    

    def end_test(self,data,result):
        tmp_color = self.define_color_result(result)      
        self.update_body_summary(result, tmp_color)
        self.update_result_html(result, tmp_color)        
        is_save_by_test = built_in.get_variable_value('${saveBYTEST}')        
        if is_save_by_test=='True':
            filename=self.get_name_of_report(data.name)
            self.print_evidence(filename)
        self.update_tcs_list(data.name,result)
        self.send_request_to_update_issues_on_jira(2,0)


    def end_suite(self,data,result):
        option = 0
        is_save_by_test = built_in.get_variable_value('${saveBYTEST}')
        if is_save_by_test=='False':
            option = 1
            filename=self.get_name_of_report(built_in.get_variable_value('${nombredocevidencia}'))                        
            self.print_evidence(filename)
            self.update_suite_summary(data.name,result,2)
        self.send_request_to_update_issues_on_jira(1,option)


    def update_tcs_list(self,tc_name,result):
        try:
            is_update_jira = ''
            #print('actualizadno informacion por test')
            try:
                is_update_jira = built_in.get_variable_value('${actualizarJira}')
            except Exception as exception:
                is_update_jira = 'false'
            if self.is_same_values(is_update_jira,'false'):
                return
            array_tc = built_in.get_variable_value('${TCSummaryList}')
            verify_exist, tc_detail = self._exist_tc_in_summary(built_in.get_variable_value('${ImagenID_TC}'),array_tc)
            if verify_exist == True:
                print('Ya existe información del TC')
                if self.is_same_values(result.status,'FAIL'):
                    tc_detail['status'] = result.status
                    tc_detail['message'] = result.message
                tmp_ejecucion = ''
                try:
                    tmp_ejecucion = built_in.get_variable_value('${EjecucionID_TC}')
                    if len(tmp_ejecucion) <= 0:
                        tmp_ejecucion = '1'
                except Exception as exception:
                    tmp_ejecucion = '1'
                list_ejecuciones = tc_detail['ejecuciones']
                detalle_ejecucion= {}
                detalle_ejecucion['ejecucion'] = tmp_ejecucion
                detalle_ejecucion['status'] = result.status
                detalle_ejecucion['message'] = result.message
                detalle_ejecucion['evidencias'] = built_in.get_variable_value('${file_arrays}')
                list_ejecuciones.append(detalle_ejecucion)
                tc_detail['ejecuciones']=list_ejecuciones
            else:
                tc_detail['name'] = tc_name
                tc_detail['description'] = result.doc
                tc_detail['status'] = result.status
                tc_detail['message'] = result.message
                tc_detail['id'] = built_in.get_variable_value('${ImagenID_TC}')
                tc_detail['key'] = built_in.get_variable_value('${Jira_Key}')
                tc_detail['starttime'] = result.starttime
                tc_detail['endtime'] = result.endtime
                tc_detail['elapsedtime'] = result.elapsedtime
                ejecuciones = []
                detalle_ejecucion= {}
                detalle_ejecucion['ejecucion'] = '1'
                detalle_ejecucion['status'] = result.status
                detalle_ejecucion['message'] = result.message
                detalle_ejecucion['evidencias'] = built_in.get_variable_value('${file_arrays}')
                ejecuciones.append(detalle_ejecucion)
                tc_detail['ejecuciones']=ejecuciones

            array_tc.append(tc_detail)
            built_in.set_suite_variable('${TCSummaryList}',array_tc)
            #print(f'listado de test: {array_tc}')
            
        except Exception as exception:
            print(f"update_tcs_list() had next error: {exception}")


    def update_suite_summary(self, suite_name, result, option):
        try:
            is_update_jira = ''
            try:
                is_update_jira = built_in.get_variable_value('${actualizarJira}')
            except Exception as exception:
                is_update_jira = 'false'
            if self.is_same_values(is_update_jira,'false'):
                return
            if option == 1:
                suite_dict = {}
                suite_dict['name'] = suite_name
                suite_dict['save_by_test'] = built_in.get_variable_value('${savebyTest}')
                suite_dict['new_execution'] = built_in.get_variable_value('${nuevaEjecucion}')
                suite_dict['keyEjecucion'] = built_in.get_variable_value('${keyEjecucion}')
                suite_dict['user'] = built_in.get_variable_value('${UsuarioEjecucionJira}')
                suite_dict['environment'] = built_in.get_variable_value('${EnvironmentExecution}')
                suite_dict['nombre_ejecucion'] = built_in.get_variable_value('${nameExecution}')
            elif option == 2:
                suite_dict = built_in.get_variable_value('${SuiteSummary}')
                suite_dict['status'] = result.status
                suite_dict['message'] = result.message
                suite_dict['description'] = result.doc
                suite_dict['evidences'] = built_in.get_variable_value('${file_arrays}')
                
            built_in.set_suite_variable('${SuiteSummary}',suite_dict)
        except Exception as exception:
            print(f"update_suite_summary() had next error: {exception}")


    def _exist_tc_in_summary(self,id_tc, array_tc):
        respuesta = False
        try:
            var_separar_tc = built_in.get_variable_value('${seprarTCRepetido}')
            if self.is_same_values(var_separar_tc,'true'):
                return respuesta, {}
            for element_tc in array_tc:
                if self.is_same_values(element_tc['id'],id_tc):
                    respuesta = True
                    print(f'Test encontrado en Summary {id_tc}')
                    return respuesta, element_tc
        except Exception as exception:
            print(f"_exist_tc_in_summary() had next error: {exception}")
        return respuesta, {}


    def send_request_to_update_issues_on_jira(self,sent_from, option):
        try:
            is_update_jira = ''
            name_tool = ''
            try:
                is_update_jira = built_in.get_variable_value('${actualizarJira}')
                name_tool = built_in.get_variable_value('${herramientaJira}')
            except Exception as exception:
                is_update_jira = 'false'
            if self.is_same_values(is_update_jira,'false'):
                return
            is_update_jira_for_test = built_in.get_variable_value('${actualizarJiraporTest}')
            if sent_from == 2: #La llamada del método proviene del end_test
                if self.is_same_values(is_update_jira_for_test,'false'):
                    return
            if sent_from == 1: #La llamada del método proviene del end_suite
                if self.is_same_values(is_update_jira_for_test,'true'):
                    return
            
            if self.is_same_values(name_tool,'xray') or self.is_same_values(name_tool,'x-ray'):
                self.update_xray(option)
            elif self.is_same_values(name_tool,'qmetry'):
                self.update_qmetry()
            else:
                console(f'No hay opción para esta herramienta: {name_tool}')
            #limpiar variables luego de actualizar jira
            built_in.set_suite_variable('${TCSummaryList}',[])
        except Exception as exception:
            print(f"send_request_to_update_issues_on_jira() had next error: {exception}")


    def update_xray(self,option):
        try:
            print("*****************************************")
            print('Actualización en xray')
            tc_summary = built_in.get_variable_value('${TCSummaryList}')
            var_suite_summary = built_in.get_variable_value('${SuiteSummary}')
            url_api = built_in.get_variable_value('${URL_Import_Jira}')
            var_token = built_in.get_variable_value('${token_Jira}')
            new_key = update_test_result(url_api, var_token, built_in.get_variable_value('${OUTPUTDIR}'),tc_summary,var_suite_summary)
            if option == 1:
                url_api = built_in.get_variable_value('${URL_Attach_Jira}')
                url_api = url_api.format(key=new_key)
                var_attlasian_token = built_in.get_variable_value('${AttlasianToken}')
                update_execution_result(url_api, var_token, built_in.get_variable_value('${OUTPUTDIR}'),var_suite_summary,var_attlasian_token, new_key)
            print("*****************************************")
        except Exception as exception:
            print(f"send_request_to_update_issues_on_jira() had next error: {exception}")


    def update_qmetry(self):
        try:
            print('qmetry')
            outputdir = built_in.get_variable_value('${OUTPUTDIR}')
            file_array=built_in.get_variable_value('${file_arrays}')
            
            print(outputdir)
            #update_test_result(built_in.get_variable_value('${OUTPUTDIR}'),file_array,'')
        except Exception as exception:
            print(f"send_request_to_update_issues_on_jira() had next error: {exception}")


    def update_body_summary(self,result, color):
        try:
            var_tcs = built_in.get_variable_value('${tablaResumenBodyPDF}')
            var_num = int(built_in.get_variable_value('${numTCSummary}'))
            var_num = var_num + 1
            built_in.set_suite_variable('${numTCSummary}',var_num)
            #console(f'resumen antes: {var_tcs}.')
            var_id = built_in.get_variable_value('${ImagenID_TC}')
            try:
                tmp_ejecucion = built_in.get_variable_value('${EjecucionID_TC}')
                if len(tmp_ejecucion) > 0:
                    var_id = var_id + '&nbsp;(&nbsp;' + tmp_ejecucion + '&nbsp;)'
            except Exception as exception:
                var_id = var_id
            var_time = self.time_execution(result.elapsedtime)
            var_status = result.status
            tmp_tcs = built_in.get_variable_value('${tdTablaResumenPDF}')
            var_tcs = var_tcs + tmp_tcs.format(no=var_num,id=var_id,ti=var_time,st=var_status, col=color)
            built_in.set_suite_variable('${tablaResumenBodyPDF}',var_tcs)
            #console(f'resumen: {var_tcs}.')
        except Exception as exception:
            print(f"update_body_summary() had next error: {exception}")
    

    def update_result_html(self,result, color):
        try:
            #var_body =  list()
            array_body = built_in.get_variable_value('${bodyHTML2}')
            body_test_dict = array_body[-1]
            var_body_new = built_in.get_variable_value('${resultBody}')
            now = datetime.datetime.now()
            var_body_new = var_body_new.format(doc=result.doc, tim=now.strftime("%H:%M:%S"), col=color, res=result.status, mes = result.message)
            body_test_dict['Resultado'] = var_body_new
            #console(f'update_body_html(): {array_body}')
            built_in.set_suite_variable('${bodyHTML2}',array_body)
        except Exception as exception:
            print(f"update_body_html() had next error: {exception}")


    def print_evidence(self, name):
        try:
            self.print_evidence_pdf(name)
        except Exception as exception:
            print(f"print_evidence() had next error: {exception}")
    

    def print_evidence_pdf(self, name):
        try:
            str_header = self.create_header()
            #aqui una variable global que sea como array, con tres cosas: nombre del archivo,
                
            file_dict={}            
            file_dict['filename']=name+'.pdf'
            file_dict['contentType']='application/pdf'
            file_dict['data']=''            
            #self.file_array.append(file_dict)   
            file_array=[]
            file_array.append(file_dict)
            built_in.set_suite_variable('${file_arrays}',file_array)
            self.convert_html_to_pdf(str_header, built_in.get_variable_value('${bodyHTML2}'), name + '.pdf')
            
        except Exception as exception:
            print(f"print_evidence_pdf() had next error: {exception}")
    
    
    def convert_html_to_pdf(self, var_header, array_body, pdf_file):
        try:
            
            #array_body = BuiltIn().get_variable_value('${bodyHTML2}')
            outputdir = built_in.get_variable_value('${OUTPUTDIR}')
            print(f'pdf file name: {pdf_file}')
            pdf = PDF(orientation="P", unit="mm",)
            #pdf = FPDF()
            pdf.set_font("helvetica", size=10)
            pdf.set_text_color(255)
            pdf.set_left_margin(22)
            pdf.set_right_margin(22)
            pdf.add_page()
            pdf.write_html(var_header)
            for test_executed in array_body:
                var_body_html = ""
                var_body_html = test_executed['Titulo'] + test_executed['Body'] + test_executed['Resultado']
                pdf.add_page()
                pdf.write_html(var_body_html)
            if len(pdf_file) <= 0:
                pdf_file="Evidencia"
                print(f'Nombre del pdf: {pdf_file}')
            path = os.path.join(outputdir, pdf_file)
            #print(f'path del PDF generado: {path}')
            pdf.output(path)
            #self.show_file_in_log_html(pdf_file,'Ver la evidencia generada en formato PDF.',1)
        except Exception as exception:
            print(f"convert_html_to_pdf() had next error: {exception}")
    

    def create_header(self):
        str_header=""
        try:
            now = datetime.datetime.now()
            var_release = built_in.get_variable_value('${release}')
            var_country = built_in.get_variable_value('${pais}')
            var_proj = built_in.get_variable_value('${proyecto}')
            var_tit = built_in.get_variable_value('${titulo}')
            var_subtit = built_in.get_variable_value('${titulo}')
            var_summary = built_in.get_variable_value('${tablaResumenBodyPDF}')
            var_sheet = ''
            try:
                var_sheet=built_in.get_variable_value('${hoja}')
                if not var_sheet:
                    var_sheet = built_in.get_variable_value('${pais}')
            except Exception as e:
                try:
                    var_sheet = built_in.get_variable_value('${pais}')
                except Exception as e:
                    var_sheet = ''
            tmp_presentation_html = built_in.get_variable_value('${presentacionEvidPDF}')
            str_header = tmp_presentation_html.format(fe=now.strftime("%A, %d-%B-%Y"),pa=var_country,ti=var_tit,ho=var_sheet, re=var_release, pr=var_proj, st=var_subtit, br = var_summary)
        except Exception as exception:
            print(f"create_header() had next error: {exception}")
        return str_header
    

    def get_name_of_report(self, str_name):
        try:
            if len(str(str_name))==0:
                str_name = 'DocEvidencia'
            var_formato_fecha = built_in.get_variable_value('${formatoFecha}')
            now = datetime.datetime.now()
            time_global = now.strftime(var_formato_fecha)
            #str_name= built_in.get_variable_value('${OUTPUTDIR}') + built_in.get_variable_value('${/}') + str(str_name).replace(" ","_") + time_global
            str_name= str(str_name).replace(" ","_") + time_global
            #path = os.path.join(outputdir, var_name)
        except Exception as exception:
            print(f"get_name_of_report() had next error: {exception}")
        return str_name


    def return_new_name(self,test_name,rest: Optional[int]=0):
        try:
            var_long_fecha = built_in.get_variable_value('${longitudFecha}')
            if test_name:
                var_max_caracteres = 103
                try:
                    var_max_caracteres = int(built_in.get_variable_value('${maxCaracteres}'))-int(var_long_fecha)-5-rest
                except Exception as e:
                    print('Error al convertirlo en entero')
                test_name = test_name[0:var_max_caracteres]
            else:
                test_name="EvidenciaDefault"
        except Exception as exception:
            print(f"return_new_name() had next error: {exception}")
        return str(test_name)


    def time_execution(self, vtime):
        var_full_time = ""
        try:
            vtime_int = int(vtime)
            var_s = int((vtime_int  / 1000) % 60)
            var_m = int((vtime_int  / (1000*60))  % 60)
            var_h = int((vtime_int  / (1000*60*60))  % 24)
            var_full_time = str(var_h).rjust(2, '0') + ":" + str(var_m).rjust(2, '0') + ":" + str(var_s).rjust(2, '0')
        except Exception as exception:
            print(f"time_execution() had next error: {exception}")
        return var_full_time


    def define_color_result(self,result):
        var_result_color = 'black'
        try:
            if not result.passed:
                var_result_color = built_in.get_variable_value('${colorFail}')
            else:
                var_result_color = built_in.get_variable_value('${colorPass}')
        except Exception as exception:
            print(f"time_execution() had next error: {exception}")
        return var_result_color


    def strip_accents(self,s):
        return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn')


    @keyword('Definition functionality tested start')
    def definition_functionality_tested_start(self,functionality_name):
        try:
            array_body = built_in.get_variable_value('${bodyHTML2}')
            body_test_dict = array_body[-1]
            var_body_old = body_test_dict['Body']
            var_body_previous = ''
            var_func = built_in.get_variable_value('${tmpFunct}')
            if var_func == '1':
                var_body_previous = """<p>&nbsp;</p>
                </font>
                </section>"""
            var_body_new = """
            <h4 align="left"><font color="black"><U>{fn}</U></font></h4>
            <section>
            <font color="black">"""
            var_body_new = var_body_new.format(fn=functionality_name)
            body_test_dict['Body'] = var_body_old + var_body_previous + var_body_new
            built_in.set_suite_variable('${bodyHTML2}',array_body)
            built_in.set_suite_variable('${tmpFunct}','1')
            #console(f'array_body: {array_body}')
        except Exception as exception:
            print(f"definition_functionality_tested_start() had next error: {exception}")


    @keyword('Logging info in report')
    def logging_info_in_report(self,tipo, message_log, mostrarhora, texto_extenso: Optional[str]="",aplica_texto_extenso: Optional[str]="False"):
        try:
            array_body = built_in.get_variable_value('${bodyHTML2}')
            body_test_dict = array_body[-1]
            var_body_old = body_test_dict['Body']
            var_body_new = ""
            if mostrarhora == 'True':
                var_body_new="""<p>Hora:&nbsp;{h}</p>"""
                now = datetime.datetime.now()
                var_body_new = var_body_new.format(h=now.strftime("%H:%M:%S"))
            if aplica_texto_extenso == 'True':
                var_body_new = var_body_new + """<p>{t}: {m}</p><br/>
                <section><code><font size="6">{ml}</font></code></section><p>&nbsp;</p>"""
                var_body_new = var_body_new.format(t=tipo,m=message_log,ml=texto_extenso)
            else:
                var_body_new = var_body_new + """<p>{t}:&nbsp;{m}</p>"""
                var_body_new = var_body_new.format(t=tipo,m=message_log)
            body_test_dict['Body'] = var_body_old + var_body_new
            built_in.set_suite_variable('${bodyHTML2}',array_body)
        except Exception as exception:
            print(f"logging_info_in_report() had next error: {exception}")


    @keyword('Logging evidence in report')
    def logging_evidence_in_report(self,filename_picture):
        try:
            array_body = built_in.get_variable_value('${bodyHTML2}')
            body_test_dict = array_body[-1]
            var_body_old = body_test_dict['Body']
            if len(str(filename_picture))==0 or str(filename_picture) == 'None':
                var_body_new = """<p>Hora: {h}</p><p>--</p>"""
            else:
                var_body_new = """<p>Hora: {h}</p>
                <img src="{f}" alt="Evidence" width="480" height="302">"""
            now = datetime.datetime.now()
            var_body_new = var_body_new.format(h=now.strftime("%H:%M:%S"),f=filename_picture)
            body_test_dict['Body'] = var_body_old + var_body_new
            built_in.set_suite_variable('${bodyHTML2}',array_body)
        except Exception as exception:
            print(f"Logging info in report() had next error: {exception}")


    @keyword('Logging info and evidence in report')
    def logging_info_and_evidence_in_report(self,tipo, message_log, filename_picture):
        try:
            array_body = built_in.get_variable_value('${bodyHTML2}')
            body_test_dict = array_body[-1]
            var_body_old = body_test_dict['Body']
            if len(str(filename_picture))==0 or str(filename_picture) == 'None':
                var_body_new = """<p>Hora: {h}&nbsp;&nbsp;-&nbsp;&nbsp;&nbsp;{t}: {m}</p><p>--</p>"""
            else:
                var_body_new = """<p>Hora: {h}&nbsp;&nbsp;-&nbsp;&nbsp;&nbsp;{t}: {m}</p>
                <img src="{f}" alt="Evidence" width="480" height="302">"""
            now = datetime.datetime.now()
            var_body_new = var_body_new.format(h=now.strftime("%H:%M:%S"),t=tipo,m=message_log,f=filename_picture)
            body_test_dict['Body'] = var_body_old + var_body_new
            built_in.set_suite_variable('${bodyHTML2}',array_body)
        except Exception as exception:
            print(f"Logging info in report() had next error: {exception}")


    @keyword('Logging request')
    def logging_request(self,subtitulo,cad_request):
        try:
            array_body = built_in.get_variable_value('${bodyHTML2}')
            body_test_dict = array_body[-1]
            var_body_old = body_test_dict['Body']
            var_body_new = """<section>
                <p>Hora:&nbsp;{h}</p>
                <p align="left"><font color="black"> <em>{SubT} </em></font></p>
                <p align="left"><font color="black">{CadRes}</font></p>        
            </section>"""
            now = datetime.datetime.now()
            var_body_new = var_body_new.format(h=now.strftime("%H:%M:%S"), SubT=subtitulo,CadRes=cad_request)
            body_test_dict['Body'] = var_body_old + var_body_new
            built_in.set_suite_variable('${bodyHTML2}',array_body)
        except Exception as exception:
            print(f"Logging info in report() had next error: {exception}")


    @keyword('Logging info table in report')
    def logging_info_table_in_report(self,tipo, message_log, info_tabla,mostrar_message: Optional[str]="True"):
        try:
            array_body = built_in.get_variable_value('${bodyHTML2}')
            body_test_dict = array_body[-1]
            var_body_old = body_test_dict['Body']
            var_body_new = ''
            if mostrar_message == 'False':
                var_body_new="""{tb}"""
            else:
                var_body_new="""<p>Hora: {h}&nbsp;&nbsp;-&nbsp;&nbsp;{t}: {m}</p>
                {tb}"""
            now = datetime.datetime.now()
            var_body_new = var_body_new.format(h=now.strftime("%H:%M:%S"), t=tipo, m=message_log, tb=info_tabla)
            body_test_dict['Body'] = var_body_old + var_body_new
            built_in.set_suite_variable('${bodyHTML2}',array_body)
        except Exception as exception:
            print(f"Logging info in report() had next error: {exception}")


    @keyword('Definition functionality tested end')
    def definition_functionality_tested_end(self):
        try:
            var_func = built_in.get_variable_value('${tmpFunct}')
            if var_func == '1':
                array_body = built_in.get_variable_value('${bodyHTML2}')
                body_test_dict = array_body[-1]
                var_body_old = body_test_dict['Body']
                var_body_new = """<p>&nbsp;</p>
                </font>
                </section>"""
                body_test_dict['Body'] = var_body_old + var_body_new
                built_in.set_suite_variable('${bodyHTML2}',array_body)
                built_in.set_suite_variable('${tmpFunct}','0')
        except Exception as exception:
            print(f"Logging info in report() had next error: {exception}")


    @keyword('Mostrar Evidencia Generada',tags=['Evidencia'])
    def show_file_in_log_html(self,filename, message, opcion:Optional[int]=0):
        """Imprime el enlace para el documento de Evidencia
        Tag: Evidencia
        """
        if opcion == 1:
            print('*HTML* <div id="report-or-log-link"><a href="{a}">{m}</a></div>'.format(a=filename, m=message))

    @keyword('Mostrar Contrato Guardado',tags=['Evidencia'])
    def printContrato(self, texto,patharchivo):
        """Imprime el enlace para el documento de Evidencia
        
        Tag: Evidencia
        """
        print('*HTML* <div id="report-or-log-link"><a href="{a}">{t}</a></div>'.format(a=patharchivo,t=texto))


class PDF(FPDF, HTMLMixin):

    def header(self):
        # Rendering logo:
        try:
            var_path_img = BuiltIn().get_variable_value('${pathLogo}')
            self.image(var_path_img, 20, 8, 13)
        except Exception as exception:
            print(f'Error al obtener la imagen para el PDF')
        # Setting font: helvetica bold 15
        self.set_font("helvetica", "B", 10)
        # Moving cursor to the right:
        self.cell(135)
        # Printing title:
        var_titulo = ''
        try:
            var_titulo = BuiltIn().get_variable_value('${textoEncabezado}')
        except Exception as exception:
            print(f"Error al obtener el titulo: {exception}")
            var_titulo = 'Evidencia de Pruebas'
        self.cell(txt=var_titulo, align="R")
        # Performing a line break:
        self.ln(20)
  
    def footer(self):
        # Position cursor at 1.5 cm from bottom:
        self.set_y(-15)
        # Setting font: helvetica italic 8
        self.set_font("helvetica", "I", 8)
        # Printing page number:
        self.cell(0, 5, f"Página {self.page_no()}/{{nb}}", align="C")
  
    def basic_table(self, headings, rows):
        for heading in headings:
            self.cell(40, 7, heading, 1)
        self.ln()
        for row in rows:
            for col in row:
                self.cell(40, 6, col, 1)
        self.ln()
    def improved_table(self, headings, rows, col_widths=(42, 39, 35, 40)):
        for col_width, heading in zip(col_widths, headings):
            self.cell(col_width, 7, heading, border=1, align="C")
        self.ln()
        for row in rows:
            self.cell(col_widths[0], 6, row[0], border="LR")
            self.cell(col_widths[1], 6, row[1], border="LR")
            self.cell(col_widths[2], 6, row[2], border="LR", align="R")
            self.cell(col_widths[3], 6, row[3], border="LR", align="R")
            self.ln()
        # Closure line:
        self.cell(sum(col_widths), 0, "", border="T")
 
    def colored_table(self, headings, rows, col_widths=(42, 39, 35, 42)):
        # Colors, line width and bold font:
        self.set_fill_color(255, 100, 0)
        self.set_text_color(255)
        self.set_draw_color(255, 0, 0)
        self.set_line_width(0.3)
        self.set_font(style="B")
        for col_width, heading in zip(col_widths, headings):
            self.cell(col_width, 7, heading, border=1, align="C", fill=True)
        self.ln()
        # Color and font restoration:
        self.set_fill_color(224, 235, 255)
        self.set_text_color(0)
        self.set_font()
        fill = False
        for row in rows:
            self.cell(col_widths[0], 6, row[0], border="LR", align="L", fill=fill)
            self.cell(col_widths[1], 6, row[1], border="LR", align="L", fill=fill)
            self.cell(col_widths[2], 6, row[2], border="LR", align="R", fill=fill)
            self.cell(col_widths[3], 6, row[3], border="LR", align="R", fill=fill)
            self.ln()
            fill = not fill
        self.cell(sum(col_widths), 0, "", "T")
  
    pass



